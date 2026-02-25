import os.path
from concurrent.futures import ThreadPoolExecutor
from typing import List
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from tqdm import tqdm

from app import config
from app.data_storage.milvus_operation import MilvusOperator
from app.llm_models.schema import ChunkModel
from app.service.embedding.embedding import EmbeddingModel
from app.utils.log_tools import logger
from app.utils.snowflake import generate_unique_id


class StructuredFileProcessor:
    def __init__(self, file_path: str, embedding_model: EmbeddingModel, collection_name: str):
        self.file_path = file_path
        self.file_name = os.path.basename(self.file_path)
        self.embedding_model = embedding_model
        self.conn = MilvusOperator(collection_name=collection_name)

    def expand_merged_cells_xlsx(self, file_path: str, sheet_name: str) -> pd.DataFrame:
        """
        读取 Excel 指定 sheet，若存在合并单元格则展开，
        将左上角的值填充到整个合并区域
        """
        wb = load_workbook(file_path, data_only=True)
        ws = wb[sheet_name]

        # 先转成 DataFrame
        data = ws.values
        columns = next(data)
        df = pd.DataFrame(data, columns=columns)

        # 如果没有合并单元格，直接返回
        if not ws.merged_cells.ranges:
            return df

        # 展开合并单元格
        for merged_range in ws.merged_cells.ranges:
            min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
            value = ws.cell(row=min_row, column=min_col).value

            for row in range(min_row - 2, max_row - 1):  # -2 是因为 header
                for col in range(min_col - 1, max_col):
                    df.iat[row, col] = value

        return df

    def load_and_split(self) :
        """
        加载结构化文件,并作切分

        """
        if not os.path.basename(self.file_path).endswith(".xlsx"):
            raise ValueError("Only .xlsx files are supported")
        # 使用 openpyxl 判断 & 展开合并单元格
        wb = load_workbook(self.file_path)
        sheet_names = wb.sheetnames

        for sheet_name in sheet_names:
            df = self.expand_merged_cells_xlsx(self.file_path, sheet_name)
            logger.info(f"正在处理 {sheet_name}")
            df["sheet_name"] = sheet_name
            for index, row in df.iterrows():
                row_dict = row.to_dict()
                chunk_text = str(row_dict)
                yield ChunkModel(
                    chunk_id=generate_unique_id(),
                    chunk_text=str(chunk_text),
                    file_name=self.file_name,
                    dense_vector=self.embedding_model.embed_query(chunk_text)

                )

    def insert_rows(self, chunks: List[ChunkModel], max_workers=30):
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            # 使用 map 并行执行 insert_entities，并显示进度条
            list(tqdm(executor.map(lambda chunk: self.conn.insert_entities([chunk]), chunks),
                      desc="插入数据中", total=len(chunks)))
        logger.info(f"向数据库中插入了 {len(chunks)} 行数据")






if __name__ == "__main__":
    file_path = "/Users/emilyguo/Desktop/Onboarding资料.xlsx"
    collection_name = "TestInfo_4096"
    embedding_model = EmbeddingModel(**config.embedding_config)
    processor = StructuredFileProcessor(file_path=file_path, embedding_model=embedding_model, collection_name=collection_name)
    chunks= list(processor.load_and_split())
    processor.insert_rows(chunks)

